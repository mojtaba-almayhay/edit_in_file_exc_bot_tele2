from telebot import TeleBot
from telebot import types
from openpyxl import load_workbook
from telebot.types import InlineKeyboardMarkup

#=========Basic Require
token     = ''
users_id  = []
path_file = 'data.xlsx'
#==========

workbook = load_workbook(path_file)
sheet = workbook.active
bot = TeleBot(token, num_threads=20, skip_pending=True)

@bot.message_handler(commands=['start'])
def start(message):
  if message.chat.id not in users_id:
    bot.reply_to(message,f"غير مؤهل للاستخدام ارسل ايديك للمطور\nID Your : {message.chat.id} \n===============\nDEV : @knk_1k")
  else:
    key = InlineKeyboardMarkup(row_width=2)
    a1 = types.InlineKeyboardButton(text="البحث", callback_data="ar")
    a2 = types.InlineKeyboardButton(text="الاضافة", callback_data="an")
    a3 = types.InlineKeyboardButton(text="تحميل ملف", callback_data="do")
    key.add(a1, a2)
    key.add(a3)
    bot.reply_to(message,f'This project was done by Eng. Mojtaba Al-Mayahi \nID : @knk_1k',reply_markup=key)


@bot.callback_query_handler(func=lambda m: True)
def qu(call):
  if call.data == 'ar':
    bot.edit_message_text(chat_id=call.message.chat.id,message_id=call.message.id,text=' يرجى إدخال القيمة المراد البحث عنها')
    bot.register_next_step_handler(call.message, search1)

  if call.data == 'an':
    bot.edit_message_text(chat_id=call.message.chat.id,message_id=call.message.id,text='يرجى ارسا البيانات بهكذا نمط \n Red|222|333|444')
    bot.register_next_step_handler(call.message, search2)
    
  if call.data == 'do':
    with open("data.xlsx", 'rb') as f:
      bot.send_document(call.message.chat.id,f)
      


def search1(message):
  if "/start" == message.text:
    start(message)

  else:
    try:
      mesage = float(message.text)
    except Exception as rt:
      if "could not convert string to float" in str(rt):
        bot.reply_to(message, "القيمة غير صحيحة يرجى ارسال بشكل 0.000")
      else:
        bot.reply_to(message, f" 1 حدث حطأ {rt}")
    bot.reply_to(message, "جار البحث ....")
    

    for i in sheet.values:
      if mesage == i[1]:
        if list(sheet.values).index(i) == 0:
          bot.send_message(message.chat.id, f"{list(sheet.values)[list(sheet.values).index(i)]}\n{list(sheet.values)[list(sheet.values).index(i)+1]}")
        elif len(list(sheet.values)) == list(sheet.values).index(i)+1:
          bot.send_message(message.chat.id, f"{list(sheet.values)[list(sheet.values).index(i)-1]}\n{list(sheet.values)[list(sheet.values).index(i)]}")
        else:
          bot.send_message(message.chat.id, f"{list(sheet.values)[list(sheet.values).index(i)-1]}\n{list(sheet.values)[list(sheet.values).index(i)]}\n{list(sheet.values)[list(sheet.values).index(i)+1]}")
          
    bot.send_message(message.chat.id, f"تم الانتهاء من عملية البحث")
    

def search2(message):
  if "/start" == message.text:
    start(message)

  else:
    try:
      textt = message.text
      last_row = sheet.max_row + 1
      sheet.cell(row=last_row, column=1).value = textt.split("|")[0]
      sheet.cell(row=last_row, column=2).value = float(textt.split("|")[1])
      sheet.cell(row=last_row, column=3).value = float(textt.split("|")[2])
      sheet.cell(row=last_row, column=4).value = float(textt.split("|")[3])
      workbook.save("data.xlsx")
      bot.reply_to(message, "تم حفظ البيانات في الجدول \n/start")
    except Exception as r:
      if 'list index out of range' in str(r):
        bot.reply_to(message,"يرجى ارسال البيانات بهكذا نمط \n Red|222|333|444 \n/start")
      else:
        bot.reply_to(message, f"حدث خطأ {r} 2 ")
        
        
print("Starting ...")
bot.infinity_polling()
